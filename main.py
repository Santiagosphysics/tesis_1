from flask import Flask, render_template, request, redirect, url_for, session, Response, send_file
from test import change_num, df_show
from model_2 import prepro_img, letter_pred
from werkzeug.utils import secure_filename
from PIL import Image
from io import StringIO
from io import BytesIO
import pandas as pd
import os 
import zipfile


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


app = Flask(__name__, static_url_path='/static')
app.secret_key = "12345"

@app.route('/')
def home():
    return render_template('home.html')




@app.route('/dev_df', methods=['GET', 'POST'])
def dev_df():
    if request.method == 'POST':
        num_q = int(request.form.get('numQuestions'))
        num_o = int(request.form.get('numOptions'))
        
        # Replace with your actual logic
        options = change_num(number_options=num_o, number_questions=num_q)
        df = df_show(num_options=num_q, num_questions=num_o)

        # Store the DataFrame and other data in the session
        session['df'] = df.to_html()
        session['num_q'] = num_q
        session['num_o'] = num_o
        session['options'] = options
        
        return redirect(url_for('dev_df'))
    else:
        df_html = session.get('df', "")
        num_q = session.get('num_q', 0)
        num_o = session.get('num_o', 0)
        options = session.get('options', [])
    
    return render_template('dev_df.html', df_html=df_html, num_q=num_q, num_o=num_o, options=options)



@app.route('/download_excel')
def download_excel():
    num_q = session.get('num_q', 0)
    num_o = session.get('num_o', 0)
    df = df_show(num_options=num_o, num_questions=num_q)
    
    output = BytesIO()
    wb = Workbook()
    ws = wb.active

    num_options_list = ['' for _ in range(num_q)]
    num_questions_list = [1 + i for i in range(num_q)]
    df_data = {
        'P  R  E  G  U  N  T  A': num_questions_list,
        'R  E  S  P  U  E  S  T  A ': num_options_list
    }
    df = pd.DataFrame(df_data)
    df = df.transpose()

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name='Tahoma', size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')  
            cell.border = border_style  

    first_col_letter = ws.cell(row=1, column=1).column_letter
    ws.column_dimensions[first_col_letter].width = 40  

    default_width = 10  
    for c_idx, col in enumerate(df.columns, start=2):  
        col_letter = chr(64 + c_idx)  
        ws.column_dimensions[col_letter].width = default_width

    for row in ws.rows:
        row_number = row[0].row
        ws.row_dimensions[row_number].height = 30 

    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment;filename=output.xlsx"}
    )






@app.route('/submit_answers', methods=['POST'])
def submit_answers():
    num_q = int(request.form.get('num_q'))

    options = session.get('options', [])

    df = pd.DataFrame({
        'Pregunta': [ i+1 for i in range(num_q)],
        'Respuesta': [request.form.get(f'{i}') for i in range(num_q)]
    })
    print(df)
    session['df'] = df.to_html()

    answers_student = [request.form.get(f'{i}') for i in range(num_q)]
    session['answer_student']=answers_student

    # print(answers_student)
    return redirect(url_for('dev_df'))





@app.route('/take', methods=['GET', 'POST'])
def take():
    df_html = session.get('df', None)
    return render_template('take.html', df_html=df_html)


@app.route('/prediction', methods=['GET', 'POST'])
def prediction():
    if request.method == 'POST':
        if 'imageFile' not in request.files:
            return redirect(request.url)

        file = request.files['imageFile']

        if file.filename == '':
            return redirect(request.url)

        upload_folder = './static/uploads'       
        file_path = os.path.join(upload_folder, file.filename)
    
        file_path = file_path.replace('\\', '/')
        file.save(file_path)


        upload_folder_2 = '/uploads'       
        file_path_2 = os.path.join(upload_folder_2, file.filename)
        file_path_2 = file_path_2.replace('\\', '/')    

        gray, img = prepro_img(file_path)

        answer_str = letter_pred(img, gray)

        answer_list = [i for i in answer_str]

        
        answer_student = session.get('answer_student', [])

        print('Answer list', answer_list, len(answer_list))
        print('Answer student', answer_student, len(answer_student))

        grade_point = 100/len(answer_student)
        exam_grade = 0
        for i in range(len(answer_student)):
            if answer_student[i] == answer_list[i]:
                exam_grade += grade_point



        print(answer_list)
        print(answer_student, 'respuesta del estudiante')
        # print(exam_grade)
        


        return render_template('prediction_result.html', image_path=file_path_2, answer=answer_str, exam_grade=int(exam_grade))

    return render_template('prediction.html')


















app.config['UPLOAD_FOLDER'] = './static/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'zip'}
app.config['ALLOWED_IMAGE_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

@app.route('/group_predictions', methods=['GET', 'POST'])
def group_predictions():
    if request.method == 'POST':
        if 'imageFile' not in request.files:
            return redirect(request.url)
        
        file = request.files['imageFile']

        if file.filename == '':
            return redirect(request.url)
        
        if not allowed_file(file.filename, app.config['ALLOWED_EXTENSIONS']):
            return redirect(request.url)  # Or return an error message

        # Save the uploaded zip file
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        # Unzip the file
        extracted_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'extracted')
        os.makedirs(extracted_folder, exist_ok=True)

        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(extracted_folder)

        # Process each image file
        results = []
        for root, _, files in os.walk(extracted_folder):
            for file in files:
                if not allowed_file(file, app.config['ALLOWED_IMAGE_EXTENSIONS']):
                    continue  # Skip non-image files

                image_path = os.path.join(root, file)

                # Apply existing single image prediction logic
                gray, img = prepro_img(image_path)
                answer_str = letter_pred(img, gray)
                answer_list = [i for i in answer_str]

                answer_student = session.get('answer_student', [])

                grade_point = 100 / len(answer_student) if len(answer_student) > 0 else 0
                exam_grade = 0
                for i in range(len(answer_student)):
                    if answer_student[i] == answer_list[i]:
                        exam_grade += grade_point

                results.append({
                    'image_path': image_path,
                    'answer': answer_str,
                    'exam_grade': int(exam_grade)
                })

        # Optionally, save results to a file or database
        results_file = os.path.join(app.config['UPLOAD_FOLDER'], 'results.txt')
        with open(results_file, 'w') as f:
            for result in results:
                f.write(f"Image: {result['image_path']}\n")
                f.write(f"Answer: {result['answer']}\n")
                f.write(f"Grade: {result['exam_grade']}\n\n")

        return send_file(results_file, as_attachment=True)  # Provide the results file for download

    return render_template('group_predictions.html')  # Render your form for file upload







if __name__ == '__main__':
    app.run(debug = True)


