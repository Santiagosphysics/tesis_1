import numpy as np 
import pandas as pd 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def change_num(number_options, number_questions):
    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    options =  [abc[i] for i in range(number_options)]

    return options



def creation_test(number_questions, number_options, options):
    df = { i+1:[ 'O' for _ in range(number_options)] for i in range(number_questions)}
    df = pd.DataFrame(df)
    options = change_num(number_options, number_questions)
    df.index = options
    return df


# def df_show(num_options, num_questions):
#     num_options_list = ['' for i in range(num_options)]
#     num_questions_list = [1+i for i in range(num_options)]

#     df = {
#         'P R E G U N T A': num_questions_list,
#         'R E S P U E S T A ': num_options_list
#         }
    
#     df = pd.DataFrame(df)
#     df = df.transpose()
#     # df = df.to_csv('df.csv', header=False)
#     return df



def df_show(num_options, num_questions):
    num_options_list = ['' for _ in range(num_questions)]
    num_questions_list = [1 + i for i in range(num_questions)]

    df = {
        'P  R  E  G  U  N  T  A': num_questions_list,
        'R  E  S  P  U  E  S  T  A ': num_options_list
    }
    df = pd.DataFrame(df)
    df = df.transpose()

    wb = Workbook()
    ws = wb.active

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name='Tahoma', size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')  
            cell.border = border_style  

    first_col_letter = ws.cell(row=1, column=1).column_letter
    ws.column_dimensions[first_col_letter].width = 40  

    default_width = 10  
    for c_idx, col in enumerate(df.columns, start=2):  
        col_letter = chr(64 + c_idx)  
        ws.column_dimensions[col_letter].width = default_width

    for row in ws.rows:
        row_number = row[0].row
        ws.row_dimensions[row_number].height = 30 

    return df